# 101, 104
import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

# 기본 폴더 경로
input_folder = r'C:\Users\USER\Desktop\기획\프로젝트\pdf\basic_folder'
output_folder = r'C:\Users\USER\Desktop\기획\프로젝트\pdf\check_noti_output_folder'

# 오늘 날짜 및 3개월 전 날짜
today = datetime.today()
three_months_ago = today - timedelta(days=90)

# 폴더 생성 (출력 폴더가 없을 경우 생성)
os.makedirs(output_folder, exist_ok=True)

# 파일 읽기 및 처리
for file_name in os.listdir(input_folder):
    if file_name.endswith('.xlsx') and 'basic' in file_name:
        file_path = os.path.join(input_folder, file_name)
        
        # Excel 파일 읽기
        df = pd.read_excel(file_path)

        # '진료시작일'에서 시분초 제거 및 날짜만 남기기
        df['진료시작일'] = pd.to_datetime(df['진료시작일']).dt.date

        # 조건 1: 진료시작일이 3개월 이내이고, 진단과가 '일반의'가 아니며, 주상병 코드가 '$'가 아닐 때
        condition_1 = (df['진료시작일'] >= three_months_ago.date()) & \
                      (df['진단과'] != '일반의') & \
                      (df['주상병\n코드'] != '$')

        # 조건 1을 만족하는 경우가 없으면 조건 2를 평가
        if not df[condition_1].empty:
            # 조건 2: 진료시작일이 3개월 이내이고, 입원/외래가 '입원'일 때
            condition_2 = (df['진료시작일'] >= three_months_ago.date()) & \
                          (df['입원/\n외래'] == '입원')

            # 조건을 만족하는 행들만 추출
            filtered_df1 = df[condition_1]
            filtered_df2 = df[condition_2]
        else:
            # 조건 1을 만족하지 않는 경우에만 조건 2 평가
            condition_2 = (df['진료시작일'] >= three_months_ago.date()) & \
                          (df['입원/\n외래'] == '입원')

            # 조건을 만족하는 행에 상태 필드 추가 (104)
            df.loc[condition_2, '상태'] = '입원(104)'

            # 조건을 만족하는 행들만 추출
            filtered_df1 = pd.DataFrame()  # 빈 DataFrame 생성
            filtered_df2 = df[condition_2]

        # 파일 이름에서 ID 추출
        id_code = file_name.split('_')[0]
        date = file_name.split('_')[2]
        time = file_name.split('_')[3]
        
        # 상태 필드 추가 및 테두리 설정을 위한 코드 공통화
        def add_status_column(ws, df, status_value):
            # 상태 필드 추가
            status_col = len(ws[1]) + 1  # 상태 필드 열 인덱스 정의
            ws.cell(row=1, column=status_col, value='상태').font = Font(bold=True)  # 상태 열 글씨를 볼드 처리
            
            # 가운데 정렬
            ws.cell(row=1, column=status_col).alignment = Alignment(horizontal='center')

            for i in range(2, len(df) + 2):  # 데이터 행 수 만큼 반복
                ws.cell(row=i, column=status_col, value=status_value)  # 상태 값 추가
            
            # 상태 필드에만 테두리 추가 (헤더 행에만)
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            ws.cell(row=1, column=status_col).border = thin_border  # 헤더 셀에만 테두리 추가

        if not filtered_df1.empty:
            # 새로운 파일 이름 생성
            output_file_name1 = f"{id_code}_101_{date}_{time}.xlsx"
            output_file_path1 = os.path.join(output_folder, output_file_name1)

            # 결과를 Excel로 저장
            filtered_df1.to_excel(output_file_path1, index=False)

            # 엑셀 파일 열기 및 B열 너비 조정
            wb = load_workbook(output_file_path1)
            ws = wb.active
            ws.column_dimensions['B'].width = 12  # B열의 너비를 12로 설정

            # 상태 열 추가 및 테두리 설정
            add_status_column(ws, filtered_df1, '질병확정진단(101)')

            # 변경 사항 저장
            wb.save(output_file_path1)
            print(f"{output_file_name1} 파일이 생성되었습니다.")
        else:
            print(f"{id_code}님은 최근 3개월 이내 질병확정진단(1-1)에 조건을 만족하는 데이터가 없습니다.")

        if not filtered_df2.empty:
            # 새로운 파일 이름 생성
            output_file_name2 = f"{id_code}_104_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_file_path2 = os.path.join(output_folder, output_file_name2)

            # 결과를 Excel로 저장
            filtered_df2.to_excel(output_file_path2, index=False)

            # 엑셀 파일 열기 및 B열 너비 조정
            wb = load_workbook(output_file_path2)
            ws = wb.active
            ws.column_dimensions['B'].width = 12  # B열의 너비를 12로 설정

            # 상태 열 추가 및 테두리 설정
            add_status_column(ws, filtered_df2, '입원(104)')

            # 변경 사항 저장
            wb.save(output_file_path2)
            print(f"{output_file_name2} 파일이 생성되었습니다.")
        else:
            print(f"{id_code}님은 최근 3개월 이내 입원(1-4)에 조건을 만족하는 데이터가 없습니다.")
